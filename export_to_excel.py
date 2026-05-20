import json, sys, os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── รับ path ของ db.json จาก argument หรือใช้ค่า default ──
db_path = sys.argv[1] if len(sys.argv) > 1 else "backend/data/db.json"
out_path = sys.argv[2] if len(sys.argv) > 2 else "nutritrack_export.xlsx"

if not Path(db_path).exists():
    print(f"❌ ไม่พบไฟล์: {db_path}")
    print("   วิธีใช้: python export_to_excel.py <path/to/db.json> <output.xlsx>")
    sys.exit(1)

with open(db_path, encoding="utf-8") as f:
    db = json.load(f)

wb = Workbook()

# ── สีและ style ──
COLOR_HEADER   = "2D6A4F"   # เขียวเข้ม
COLOR_HEADER2  = "52B788"   # เขียวอ่อน
COLOR_ALT      = "F0F7F4"   # เขียวจาง (แถวสลับ)
COLOR_ACCENT   = "F4A261"   # ส้ม
COLOR_WARN     = "FDEBD0"

def hdr_font(color="FFFFFF"):
    return Font(name="Arial", bold=True, color=color, size=10)

def cell_font():
    return Font(name="Arial", size=10)

def hdr_fill(hex_color=COLOR_HEADER):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def alt_fill():
    return PatternFill("solid", start_color=COLOR_ALT, fgColor=COLOR_ALT)

def center():
    return Alignment(horizontal="center", vertical="center")

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header_row(ws, row, col_count, bg=COLOR_HEADER):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font("FFFFFF")
        cell.fill = hdr_fill(bg)
        cell.alignment = center()
        cell.border = thin_border()

def style_data_row(ws, row, col_count, alt=False):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = cell_font()
        if alt:
            cell.fill = alt_fill()
        cell.border = thin_border()

def auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        length = max((len(str(cell.value or "")) for cell in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)

def write_sheet(ws, title, headers, rows, bg=COLOR_HEADER):
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    title_cell.fill = hdr_fill(COLOR_ACCENT)
    title_cell.alignment = center()
    ws.row_dimensions[1].height = 24

    # Header row
    for ci, h in enumerate(headers, 1):
        ws.cell(row=2, column=ci, value=h)
    style_header_row(ws, 2, len(headers), bg)
    ws.row_dimensions[2].height = 20

    # Data rows
    for ri, row in enumerate(rows, 3):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
        style_data_row(ws, ri, len(headers), alt=(ri % 2 == 0))

    auto_width(ws)
    ws.freeze_panes = "A3"

# ═══════════════════════════════════════════
# Sheet 1 — บันทึกอาหาร (Food Logs)
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = "🍽️ บันทึกอาหาร"

user_map = {u["id"]: u.get("name", u.get("email", "?")) for u in db.get("users", [])}
meal_map = {"breakfast": "มื้อเช้า", "lunch": "มื้อกลางวัน", "dinner": "มื้อเย็น", "snack": "อาหารว่าง"}

logs = sorted(db.get("foodLogs", []), key=lambda x: (x.get("date",""), x.get("meal","")))
rows1 = []
for l in logs:
    rows1.append([
        l.get("date", ""),
        meal_map.get(l.get("meal", ""), l.get("meal", "")),
        user_map.get(l.get("userId", ""), "?"),
        l.get("foodName", ""),
        l.get("quantity", 0),
        l.get("calories", 0),
        round(l.get("carbs", 0), 1),
        round(l.get("protein", 0), 1),
        round(l.get("fat", 0), 1),
        round(l.get("sugar", 0), 1),
        l.get("sodium", 0),
        l.get("cholesterol", 0),
    ])

headers1 = ["วันที่", "มื้อ", "ผู้ใช้", "ชื่ออาหาร", "ปริมาณ (g)", "พลังงาน (kcal)",
            "คาร์บ (g)", "โปรตีน (g)", "ไขมัน (g)", "น้ำตาล (g)", "โซเดียม (mg)", "คอเลสเตอรอล (mg)"]
write_sheet(ws1, "บันทึกอาหารทั้งหมด", headers1, rows1)

# ═══════════════════════════════════════════
# Sheet 2 — สรุปรายวัน (Daily Summary)
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("📊 สรุปรายวัน")

from collections import defaultdict
daily = defaultdict(lambda: defaultdict(float))
for l in db.get("foodLogs", []):
    d = l.get("date", "")
    uid = l.get("userId", "")
    key = (d, uid)
    daily[key]["calories"]    += l.get("calories", 0)
    daily[key]["carbs"]       += l.get("carbs", 0)
    daily[key]["protein"]     += l.get("protein", 0)
    daily[key]["fat"]         += l.get("fat", 0)
    daily[key]["sugar"]       += l.get("sugar", 0)
    daily[key]["sodium"]      += l.get("sodium", 0)
    daily[key]["cholesterol"] += l.get("cholesterol", 0)

act_daily = defaultdict(float)
for a in db.get("activities", []):
    key = (a.get("date",""), a.get("userId",""))
    act_daily[key] += a.get("caloriesBurned", 0)

# หา goal ล่าสุดของแต่ละ user
goal_map = {}
for g in db.get("goals", []):
    goal_map[g["userId"]] = g

rows2 = []
for (date, uid), s in sorted(daily.items()):
    burned = act_daily.get((date, uid), 0)
    net    = s["calories"] - burned
    goal   = goal_map.get(uid, {})
    g_cal  = goal.get("calories", 2000)
    status = "✅ ในเป้า" if net <= g_cal else "⚠️ เกินเป้า"
    rows2.append([
        date,
        user_map.get(uid, "?"),
        round(s["calories"], 0),
        round(burned, 0),
        round(net, 0),
        g_cal,
        status,
        round(s["carbs"], 1),
        round(s["protein"], 1),
        round(s["fat"], 1),
        round(s["sugar"], 1),
        round(s["sodium"], 0),
        round(s["cholesterol"], 0),
    ])

headers2 = ["วันที่", "ผู้ใช้", "รับเข้า (kcal)", "เผาผลาญ (kcal)", "สุทธิ (kcal)", "เป้าหมาย (kcal)",
            "สถานะ", "คาร์บ (g)", "โปรตีน (g)", "ไขมัน (g)", "น้ำตาล (g)", "โซเดียม (mg)", "คอเลสเตอรอล (mg)"]
write_sheet(ws2, "สรุปโภชนาการรายวัน (หักออกกำลังกายแล้ว)", headers2, rows2, bg=COLOR_HEADER2)

# ═══════════════════════════════════════════
# Sheet 3 — กิจกรรมออกกำลังกาย
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("🏃 กิจกรรม")

int_map = {"low": "เบา", "medium": "ปานกลาง", "high": "หนัก"}
acts = sorted(db.get("activities", []), key=lambda x: x.get("date",""))
rows3 = [[
    a.get("date",""),
    user_map.get(a.get("userId",""), "?"),
    a.get("name",""),
    a.get("type",""),
    int_map.get(a.get("intensity",""), a.get("intensity","")),
    a.get("duration", 0),
    a.get("caloriesBurned", 0),
    a.get("notes",""),
] for a in acts]

headers3 = ["วันที่", "ผู้ใช้", "ชื่อกิจกรรม", "ประเภท", "ความเข้มข้น",
            "ระยะเวลา (นาที)", "เผาผลาญ (kcal)", "บันทึก"]
write_sheet(ws3, "กิจกรรมออกกำลังกาย", headers3, rows3, bg="457B9D")

# ═══════════════════════════════════════════
# Sheet 4 — คลังเมนูอาหาร
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("🥗 คลังเมนู")

foods = db.get("foodItems", [])
rows4 = []
for f in foods:
    rows4.append([
        f.get("name",""),
        "✅ เมนูของฉัน" if f.get("custom") else "📋 เมนูมาตรฐาน",
        user_map.get(f.get("userId",""), "-") if f.get("custom") else "-",
        f.get("servingSize", 100),
        f.get("unit","g"),
        f.get("calories", 0),
        f.get("carbs", 0),
        f.get("protein", 0),
        f.get("fat", 0),
        f.get("sugar", 0),
        f.get("sodium", 0),
        f.get("cholesterol", 0),
    ])

headers4 = ["ชื่อเมนู", "ประเภท", "เจ้าของ", "ปริมาณ/ครั้ง", "หน่วย",
            "พลังงาน (kcal)", "คาร์บ (g)", "โปรตีน (g)", "ไขมัน (g)",
            "น้ำตาล (g)", "โซเดียม (mg)", "คอเลสเตอรอล (mg)"]
write_sheet(ws4, "คลังเมนูอาหาร (มาตรฐาน + ของฉัน)", headers4, rows4, bg="E76F51")

# ═══════════════════════════════════════════
# Sheet 5 — เป้าหมาย
# ═══════════════════════════════════════════
ws5 = wb.create_sheet("🎯 เป้าหมาย")

goals_list = db.get("goals", [])
rows5 = [[
    user_map.get(g.get("userId",""), "?"),
    g.get("calories", 0),
    g.get("carbs", 0),
    g.get("protein", 0),
    g.get("fat", 0),
    g.get("sugar", 0),
    g.get("sodium", 0),
    g.get("cholesterol", 0),
    g.get("updatedAt","")[:10] if g.get("updatedAt") else "",
] for g in goals_list]

headers5 = ["ผู้ใช้", "พลังงาน (kcal)", "คาร์บ (g)", "โปรตีน (g)", "ไขมัน (g)",
            "น้ำตาล (g)", "โซเดียม (mg)", "คอเลสเตอรอล (mg)", "อัปเดตล่าสุด"]
write_sheet(ws5, "เป้าหมายโภชนาการ", headers5, rows5, bg="52B788")

# ── บันทึกไฟล์ ──
wb.save(out_path)
print(f"✅ Export สำเร็จ → {out_path}")
print(f"   📋 บันทึกอาหาร:  {len(rows1)} รายการ")
print(f"   📊 สรุปรายวัน:   {len(rows2)} วัน")
print(f"   🏃 กิจกรรม:      {len(rows3)} รายการ")
print(f"   🥗 คลังเมนู:     {len(rows4)} รายการ")
print(f"   🎯 เป้าหมาย:     {len(rows5)} รายการ")
